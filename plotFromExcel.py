
import math
import numpy as np
import numpy.linalg as npla

import matplotlib.pyplot as plt
from matplotlib import cm
from scipy.interpolate import make_interp_spline

%matplotlib inline
#%matplotlib tk

np.set_printoptions(precision = 4)

# Import `load_workbook` module from `openpyxl`
from openpyxl import load_workbook


def drawVelocity(pathname):
    # Load in the workbook
    wb = load_workbook(pathname)

    # Get sheet names
    print(wb.sheetnames)
    
    sheet = wb["Curves"]

    length = sheet.max_row
    #First three values in velocity are invalid, as well as last value
    time = np.zeros(length-4);
    velocity= np.zeros(length-4);
    #Loop through and assign time and velocity values
    index = 0;
    for cellObj in sheet['D4':'E'+str(sheet.max_row-1)]:
        time[index] = cellObj[0].value
        velocity[index] = cellObj[1].value
        index+=1
    #return them
    return time,velocity


def drawForce(pathname):
    # Load in the workbook
    wb = load_workbook(pathname)

    # Get sheet names
    print(wb.sheetnames)

    sheet = wb["Curves"]
    
    #Get the maximum row number
    length = sheet.max_row
    #For Force, the first two values are invalid
    time = np.zeros(length-2);
    force= np.zeros(length-2);
    #loop through and assign seconds and force total values
    index = 0;
    for cellObj in sheet['G3':'J'+str(sheet.max_row)]:
        time[index] = cellObj[0].value
        force[index] = cellObj[3].value
        index+=1
    #return the two arrays
    return time,force 

def norm(data):
    #min = np.min(data)
    #max = np.max(data)
    #return [(val - min) / (max - min) for val in data]]
    return data/ np.linalg.norm(data)

def smooth(time,y):
    x_new = np.linspace(time.min(), time.max(), 300)
    a_BSpline = make_interp_spline(time, y)
    y_new = a_BSpline(x_new)

    return x_new,y_new